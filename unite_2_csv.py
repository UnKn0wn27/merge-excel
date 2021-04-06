import csv
import json

from datetime import datetime
from dateutil.relativedelta import relativedelta
from typing import Any, Dict, List, Union

from openpyxl import Workbook


class Unite2CSV:
    def __init__(
        self,  
        csv_1: str, 
        csv_2: str, 
        csv_1_primary_field: str, 
        csv_2_primary_field: str, 
        time_series_by_column: List[str], 
        get_data_from_csv_2: List[str], 
        ignore_these_columns: List[str]
    ) -> None:
        self.date_formats = ['%Y-%m-%d', '%m/%d/%y', '%m/%d/%Y']
        self.csv_1 = self.read_data(csv_1, csv_1_primary_field)
        self.csv_2 = self.read_data(csv_2, csv_2_primary_field)
        self.primary_field_1 = csv_1_primary_field
        self.primary_field_2 = csv_2_primary_field
        self.time_series_by_column = time_series_by_column
        self.get_data_from_csv_2 = get_data_from_csv_2
        self.ignore_these_columns = ignore_these_columns

    def read_data(self, csv_file: str, field: str) -> List[dict]:
        return_data = []
        with open(csv_file) as f:
            data = csv.DictReader(f, delimiter=',', quotechar='"')

            for index, d in enumerate(data):
                if isinstance(d[field], datetime):
                    continue

                for date_format in self.date_formats:
                    try:
                        d[field] = datetime.strptime(d[field], date_format)
                        return_data.append(d)
                        break
                    except ValueError:
                        pass
                
                if not isinstance(d[field], datetime):
                    print(f"File {csv_file}. Row {index + 2}:")
                    print(d)
                    raise ValueError(f'"{d[field]}"" Field was unable to convert to date!')

            return return_data

    def what_date_to_get(self, date: datetime, amount: int, period: str) -> Union[str, datetime]:
        description, return_date = None, None
        if amount != 0:
            if period == 'day':
                return_date = date + relativedelta(days=amount)
            if period == 'month':
                return_date = date + relativedelta(months=amount)
            if period == 'year':
                return_date = date + relativedelta(years=amount)
            description = f'{amount} {period}/{period}s'
        return description, return_date

    def return_matching_fields_by_date(self, data: List[dict], primary_field: str, by_value: Any) -> List[List[dict]]:
        return_data = []
        matching_date = None
        for d in data:
            if d[primary_field] == by_value:
                matching_date = d[primary_field]
                d['date_that_was_taken'] = f"Current day/month/year."
                return_data.append(d)
                break

        if matching_date:
            period_dict = {}
            for period_to_get in self.get_data_from_csv_2:
                day, period = period_to_get.split(' ')
                column_text, date = self.what_date_to_get(
                    matching_date, int(day), period)
                period_dict[date] = column_text

            for d in data:
                if d[primary_field] in period_dict.keys():
                    d['date_that_was_taken'] = period_dict[d[primary_field]]
                    return_data.append(d)
                    del period_dict[d[primary_field]]
                elif d[primary_field].isoweekday() == 1:
                    return_date = d[primary_field] - relativedelta(days=2)
                    if return_date in period_dict.keys():
                        d['date_that_was_taken'] = period_dict[return_date]
                        return_data.append(d)
                        del period_dict[return_date]
                elif d[primary_field].isoweekday() == 5:
                    return_date = d[primary_field] + relativedelta(days=2)
                    if return_date in period_dict.keys():
                        d['date_that_was_taken'] = period_dict[return_date]
                        return_data.append(d)
                        del period_dict[return_date]
                
                if not period_dict:
                    break

        return return_data

    def group_together(self) -> List[dict]:
        return_data = []

        for csv_1_data in self.csv_1:

            primary_value = csv_1_data[self.primary_field_1]
            csv_1_data.pop(self.primary_field_1, None)

            matching_data = self.return_matching_fields_by_date(
                self.csv_2, self.primary_field_2, primary_value)

            if not matching_data:
                continue

            for original_csv_2_data in matching_data:
                csv_2_data = original_csv_2_data.copy()
                csv_2_data.pop(self.primary_field_2, None)

                group_dict = {
                    'primary_value': primary_value.strftime('%Y-%m-%d')}
                updated_csv_1_data = {f'csv_1_{k}' if k in csv_2_data.keys(
                ) else k: v for k, v in csv_1_data.items()}
                updated_csv_2_data = {f'csv_2_{k}' if k in csv_1_data.keys(
                ) else k: v for k, v in csv_2_data.items()}

                group_dict.update(updated_csv_1_data)
                group_dict.update(updated_csv_2_data)

                for key in self.ignore_these_columns:
                    if group_dict.get(key):
                        del group_dict[key]

                return_data.append(group_dict)

        return return_data

    def to_csv(self, data: List[dict], fname: str) -> None:
        if not data:
            raise ValueError('No data matched primary field!')

        fieldnames = data[0].keys()
        with open(fname, 'w') as f:
            csv_file = csv.DictWriter(f, fieldnames=fieldnames)
            csv_file.writeheader()
            csv_file.writerows(data)

    def time_series(self, data: List[dict]) -> Dict[str, Dict[str, dict]]:
        return_dict = {}
        for d in data.copy():
            primary_column = d['primary_value']
            return_dict.setdefault(primary_column, {})
            attribute = ', '.join(
                [v for k, v in d.items() if k in self.time_series_by_column])
            return_dict[primary_column].setdefault(attribute, {})

            for column in self.time_series_by_column + ['primary_value']:
                del d[column]

            date_that_was_taken = d['date_that_was_taken']
            del d['date_that_was_taken']

            return_dict[primary_column][attribute].setdefault(date_that_was_taken, {})
            
            return_dict[primary_column][attribute][date_that_was_taken] = d
        return return_dict

    def time_series_to_xlsx(self, data: Dict[str, Dict[str, dict]], fname: str) -> None:
        wb = Workbook()
        ws = wb.active

        for date_index, date in enumerate(data.keys()):
            ws.cell(column=1, row=date_index+4, value=date)

        gather_attr = []
        gather_periods = []
        gather_subtitles = []
        gather_subtitle_values = []
        nr_of_subtitles_attr = 0
        
        for items in data.values():
            for key, period_values in items.items():
                if key in gather_attr:
                    continue
                
                for period_key, values in period_values.items():
                    if period_key not in gather_periods:
                        gather_periods.append(period_key)

                    gather_subtitles.extend(values.keys())
                    gather_subtitle_values.append(list(values.values()))
                    if nr_of_subtitles_attr < len(values.keys()):
                        nr_of_subtitles_attr = len(values.keys())

                gather_attr.append(key)
                
        attr_start_column = 2
        attr_end_column = nr_of_subtitles_attr * len(gather_periods) + 1

        period_start_column = 2
        period_end_column = nr_of_subtitles_attr + 1
        for attr in gather_attr:
            ws.merge_cells(start_row=1, start_column=attr_start_column,
                           end_row=1, end_column=attr_end_column)
             
            ws.cell(column=attr_start_column, row=1, value=attr)

            attr_start_column = attr_end_column + 1
            attr_end_column = attr_start_column + (nr_of_subtitles_attr * len(gather_periods) - 1)

            for period in gather_periods:
                ws.merge_cells(start_row=2, start_column=period_start_column,
                               end_row=2, end_column=period_end_column)
                ws.cell(column=period_start_column, row=2, value=period)
                period_start_column = period_end_column + 1
                period_end_column = period_start_column + (nr_of_subtitles_attr - 1) 

        for subtitle_index, subtitle in enumerate(gather_subtitles):
            ws.cell(column=subtitle_index+2, row=3, value=subtitle)

        use_attribute = None
        use_period = None

        date_column = [c.value for c in ws['A:A'][3:]]
        attribute_row = [c.value for c in ws[1]]
        period_row = [c.value for c in ws[2]]
        subtitle_row = [c.value for c in ws[3]]
        for row_index, use_date in enumerate(date_column):
            for coll_index, attribute in enumerate(attribute_row):
                if attribute:
                    use_attribute = attribute

                if not data[use_date].get(use_attribute):
                    continue

                if period_row[coll_index]:
                    use_period = period_row[coll_index]

                if not data[use_date][use_attribute].get(use_period):
                    continue

                subtitle_value = subtitle_row[coll_index]
                if subtitle_value != None:
                    value = data[use_date][use_attribute][use_period][subtitle_value]
                    ws.cell(column=coll_index+1, row=row_index + 4, value=value)

        wb.save(filename=fname)


if __name__ == "__main__":
    import time
    start = time.time()
    unite2csv = Unite2CSV(
        csv_1='Drought3.csv',
        # csv_1='WASDE (input X) copy.csv',
        csv_2='WheatDec1.csv',
        csv_1_primary_field='Report date',
        csv_2_primary_field='Date',
        time_series_by_column=['State'],
        get_data_from_csv_2=['1 day', '-1 day'],#, '1 month', '-1 month', '1 year', '-1 year'],
        ignore_these_columns = ['', 'Ingestion timestamp', 'Market year start', 'Market year end']
    )

    gt = time.time()
    united_data = unite2csv.group_together()
    print('unite2csv', time.time() - gt)
  
    # unite2csv.to_csv(united_data, 'united_data.csv')
    tsd = time.time()
    time_series_data = unite2csv.time_series(united_data)
    print('time_series_data', time.time() - tsd)

    tstx = time.time()
    unite2csv.time_series_to_xlsx(time_series_data, 'united_data.xlsx')
    print('time_series_to_xlsx', time.time() - tstx)

    # 2010-07-09
    print("End:", time.time() - start)
